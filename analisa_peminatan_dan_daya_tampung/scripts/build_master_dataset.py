import os
import openpyxl
import pandas as pd
import numpy as np
from scipy import stats

def clean_str(val):
    if val is None or pd.isna(val):
        return ""
    return str(val).strip()

def get_fakultas_and_klaster(kd, nm, jenjang):
    nm_upper = nm.upper()
    
    # 1. PSDKU Gayo Lues
    if 'GAYO LUES' in nm_upper or 'PSDKU' in nm_upper:
        return 'PSDKU Gayo Lues', 'PSDKU Gayo Lues'
    
    # 2. Diploma 3 Vokasi
    if jenjang == 'D3' or 'D3' in nm_upper:
        if any(x in nm_upper for x in ['LISTRIK', 'MESIN', 'SIPIL']): return 'Teknik', 'Diploma 3 Vokasi'
        if any(x in nm_upper for x in ['SEKRETARI', 'PERUSAHAAN', 'AKUNTANSI', 'PERBANKAN']): return 'Ekonomi dan Bisnis', 'Diploma 3 Vokasi'
        if any(x in nm_upper for x in ['AGRIBISNIS', 'PETERNAKAN']): return 'Pertanian', 'Diploma 3 Vokasi'
        if 'HEWAN' in nm_upper: return 'Kedokteran Hewan', 'Diploma 3 Vokasi'
        if 'INFORMATIKA' in nm_upper: return 'MIPA', 'Diploma 3 Vokasi'
        return 'Vokasi', 'Diploma 3 Vokasi'
    
    # 3. S1 Kampus Utama & D4
    klaster = 'S1 Kampus Utama'
    if kd.startswith('01'): return 'Ekonomi dan Bisnis', klaster
    if kd.startswith('02'): return 'Kedokteran Hewan', klaster
    if kd.startswith('03'): return 'Hukum', klaster
    if kd.startswith('04'): return 'Teknik', klaster
    if kd.startswith('05'): return 'Pertanian', klaster
    if kd.startswith('06'): return 'FKIP', klaster
    if kd.startswith('07'): return 'Kedokteran', klaster
    if kd.startswith('08'): return 'MIPA', klaster
    if kd.startswith('10'): return 'FISIP', klaster
    if kd.startswith('11'): return 'Kelautan dan Perikanan', klaster
    if kd.startswith('12'): return 'Keperawatan', klaster
    if kd.startswith('13'): return 'Kedokteran Gigi', klaster
    
    # Fallback by name
    if 'PERMINYAKAN' in nm_upper: return 'Teknik', klaster
    if 'HUBUNGAN INTERNASIONAL' in nm_upper: return 'FISIP', klaster
    if any(x in nm_upper for x in ['HUKUM']): return 'Hukum', klaster
    if any(x in nm_upper for x in ['DOKTER GIGI']): return 'Kedokteran Gigi', klaster
    if any(x in nm_upper for x in ['DOKTER', 'PSIKOLOGI']): return 'Kedokteran', klaster
    if any(x in nm_upper for x in ['KEPERAWATAN']): return 'Keperawatan', klaster
    if any(x in nm_upper for x in ['TEKNIK', 'ARSITEKTUR', 'PERENCANAAN']): return 'Teknik', klaster
    if any(x in nm_upper for x in ['PERTANIAN', 'AGRO', 'AGRI', 'PETERNAKAN', 'TANAH', 'KEHUTANAN']): return 'Pertanian', klaster
    if any(x in nm_upper for x in ['PERIKANAN', 'KELAUTAN', 'PERAIRAN']): return 'Kelautan dan Perikanan', klaster
    if any(x in nm_upper for x in ['PENDIDIKAN', 'PGSD', 'PENJAS', 'BIMBINGAN']): return 'FKIP', 'S1 Kampus Utama'
    if any(x in nm_upper for x in ['EKONOMI', 'MANAJEMEN', 'AKUNTANSI', 'BISNIS']): return 'Ekonomi dan Bisnis', klaster
    if any(x in nm_upper for x in ['POLITIK', 'SOSIOLOGI', 'KOMUNIKASI', 'PEMERINTAHAN', 'HUBUNGAN']): return 'FISIP', klaster
    if any(x in nm_upper for x in ['MIPA', 'FISIKA', 'KIMIA', 'BIOLOGI', 'MATEMATIKA', 'INFORMATIKA', 'FARMASI', 'STATISTIKA']): return 'MIPA', klaster
    return 'Lainnya', klaster

def calculate_ols_slope(y_values):
    # y_values corresponds to [2022, 2023, 2024, 2025, 2026]
    valid_pairs = []
    x_coords = [2022, 2023, 2024, 2025, 2026]
    for x, y in zip(x_coords, y_values):
        if y > 0:
            valid_pairs.append((x, y))
    
    if len(valid_pairs) < 3:
        return 0.0, 0.0
    
    xs = [p[0] for p in valid_pairs]
    ys = [p[1] for p in valid_pairs]
    
    slope, intercept, r_val, p_val, std_err = stats.linregress(xs, ys)
    return round(slope, 2), round(r_val**2, 3)

def calculate_cagr(start_val, end_val, periods=4):
    if start_val <= 0 or end_val <= 0:
        return np.nan
    return round(((end_val / start_val) ** (1.0 / periods) - 1.0) * 100.0, 2)

def main():
    base_dir = "/Users/auliamuzhaffar/Documents/maganghub"
    file_old = os.path.join(base_dir, "tugas-5", "rekap data.xlsx")
    file_2026 = os.path.join(base_dir, "DAYA_TAMPUNG_2026_2027 GANJIL_MAGANG.xlsx")
    out_dir = os.path.join(base_dir, "tugas-5", "analisa_peminatan_dan_daya_tampung", "data")
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, "master_analisa_peminatan_dan_daya_tampung_2022_2026.xlsx")

    print(f"Loading {file_old} and {file_2026}...")
    wb_old = openpyxl.load_workbook(file_old, data_only=True)
    ws_old = wb_old["Rekapitulasi Data"]

    wb_2026 = openpyxl.load_workbook(file_2026, data_only=True)
    ws_2026 = wb_2026["DT D3_D4_S1 2026"]

    rows_old = list(ws_old.iter_rows(min_row=7, max_row=87, values_only=True))
    rows_26 = list(ws_2026.iter_rows(min_row=7, max_row=87, values_only=True))

    master_records = []
    jalur_records_all = []
    jalur_records_2022 = []
    jalur_records_2023 = []
    jalur_records_2024 = []
    jalur_records_2025 = []
    jalur_records_2026 = []

    for idx in range(81):
        r_old = rows_old[idx]
        r_26 = rows_26[idx]

        no = idx + 1
        kd_prodi = str(r_old[1] or '').strip()
        nm_prodi = str(r_old[2] or '').strip().upper()
        jenjang = str(r_old[3] or 'S1').strip().upper()

        fakultas, klaster = get_fakultas_and_klaster(kd_prodi, nm_prodi, jenjang)

        # ---------------- 2022 ----------------
        pm_22 = float(r_old[19]) if r_old[19] is not None else 0.0
        dt_22 = float(r_old[20]) if r_old[20] is not None else 0.0
        la_22 = (float(r_old[6] or 0) + float(r_old[11] or 0) + float(r_old[16] or 0))
        du_22 = (float(r_old[7] or 0) + float(r_old[12] or 0) + float(r_old[17] or 0))

        # ---------------- 2023 ----------------
        pm_23 = float(r_old[36]) if r_old[36] is not None else 0.0
        dt_23 = float(r_old[37]) if r_old[37] is not None else 0.0
        la_23 = (float(r_old[24] or 0) + float(r_old[29] or 0) + float(r_old[33] or 0))
        du_23 = (float(r_old[25] or 0) + float(r_old[30] or 0) + float(r_old[34] or 0))

        # ---------------- 2024 ----------------
        pm_24 = float(r_old[60]) if r_old[60] is not None else 0.0
        dt_24 = float(r_old[61]) if r_old[61] is not None else 0.0
        la_24 = (float(r_old[41] or 0) + float(r_old[45] or 0) + float(r_old[49] or 0) + float(r_old[53] or 0) + float(r_old[56] or 0) + float(r_old[58] or 0))
        du_24 = (float(r_old[42] or 0) + float(r_old[46] or 0) + float(r_old[50] or 0) + float(r_old[54] or 0) + float(r_old[57] or 0) + float(r_old[59] or 0))

        # ---------------- 2025 ----------------
        pm_25 = float(r_old[87]) if r_old[87] is not None else 0.0
        dt_25 = float(r_old[88]) if r_old[88] is not None else 0.0
        la_25 = (float(r_old[65] or 0) + float(r_old[69] or 0) + float(r_old[73] or 0) + float(r_old[77] or 0) + float(r_old[81] or 0) + float(r_old[84] or 0))
        du_25 = (float(r_old[66] or 0) + float(r_old[70] or 0) + float(r_old[74] or 0) + float(r_old[78] or 0) + float(r_old[82] or 0) + float(r_old[85] or 0))

        # ---------------- 2026 ----------------
        pm_26 = float(r_26[79]) if r_26[79] is not None else 0.0
        dt_26 = float(r_26[78]) if r_26[78] is not None else 0.0
        la_26 = float(r_26[80]) if r_26[80] is not None else 0.0
        du_26 = float(r_26[77]) if r_26[77] is not None else 0.0

        # Helper to construct clean pathway record
        def make_jalur_row(yr, j_std, j_asli, j_pm, j_dt, j_la, j_du):
            gugur = max(0.0, j_la - j_du)
            yield_r = round(j_du / j_la * 100.0, 1) if j_la > 0 else 0.0
            return {
                "Tahun": yr,
                "No": no,
                "Fakultas": fakultas,
                "Program_Studi": nm_prodi,
                "Jenjang": jenjang,
                "Segmen_Analisis": klaster,
                "Jalur_Penerimaan": j_std,
                "Nama_Asli_Jalur": j_asli,
                "Peminat": int(j_pm),
                "Daya_Tampung": int(j_dt),
                "Lulus_Seleksi": int(j_la),
                "Daftar_Ulang": int(j_du),
                "Tidak_Daftar_Ulang": int(gugur),
                "Yield_Rate_Persen": yield_r
            }

        # 1. Jalur 2022 (SNMPTN -> SNBP, SBMPTN -> SNBT, SMMPTN)
        j_2022_list = [
            ("SNBP", "SNMPTN (SNBP)", float(r_old[4] or 0), float(r_old[5] or 0), float(r_old[6] or 0), float(r_old[7] or 0)),
            ("SNBT", "SBMPTN (SNBT)", float(r_old[9] or 0), float(r_old[10] or 0), float(r_old[11] or 0), float(r_old[12] or 0)),
            ("SMMPTN", "SMMPTN Barat", float(r_old[14] or 0), float(r_old[15] or 0), float(r_old[16] or 0), float(r_old[17] or 0))
        ]
        for j_std, j_asli, pm, dt, la, du in j_2022_list:
            row_dict = make_jalur_row(2022, j_std, j_asli, pm, dt, la, du)
            jalur_records_2022.append(row_dict)
            jalur_records_all.append(row_dict)

        # 2. Jalur 2023 (SNBP, SNBT, SMMPTN)
        j_2023_list = [
            ("SNBP", "SNBP", float(r_old[22] or 0), float(r_old[23] or 0), float(r_old[24] or 0), float(r_old[25] or 0)),
            ("SNBT", "SNBT", float(r_old[27] or 0), float(r_old[28] or 0), float(r_old[29] or 0), float(r_old[30] or 0)),
            ("SMMPTN", "SMMPTN Barat", 0.0, float(r_old[32] or 0), float(r_old[33] or 0), float(r_old[34] or 0))
        ]
        for j_std, j_asli, pm, dt, la, du in j_2023_list:
            row_dict = make_jalur_row(2023, j_std, j_asli, pm, dt, la, du)
            jalur_records_2023.append(row_dict)
            jalur_records_all.append(row_dict)

        # 3. Jalur 2024 (SNBP, SNBT, SMMPTN, TALENTA, SMC, ADIK)
        j_2024_list = [
            ("SNBP", "SNBP", 0.0, float(r_old[40] or 0), float(r_old[41] or 0), float(r_old[42] or 0)),
            ("SNBT", "SNBT", 0.0, float(r_old[44] or 0), float(r_old[45] or 0), float(r_old[46] or 0)),
            ("SMMPTN", "SMMPTN Barat", 0.0, float(r_old[48] or 0), float(r_old[49] or 0), float(r_old[50] or 0)),
            ("TALENTA", "TALENTA USK", 0.0, float(r_old[52] or 0), float(r_old[53] or 0), float(r_old[54] or 0)),
            ("SMC", "SELEKSI MANDIRI CADANGAN", 0.0, 0.0, float(r_old[56] or 0), float(r_old[57] or 0)),
            ("ADIK", "ADIK (Afirmasi)", 0.0, 0.0, float(r_old[58] or 0), float(r_old[59] or 0))
        ]
        for j_std, j_asli, pm, dt, la, du in j_2024_list:
            row_dict = make_jalur_row(2024, j_std, j_asli, pm, dt, la, du)
            jalur_records_2024.append(row_dict)
            jalur_records_all.append(row_dict)

        # 4. Jalur 2025 (SNBP, SNBT, SMMPTN, TALENTA, SMC, ADIK)
        j_2025_list = [
            ("SNBP", "SNBP", 0.0, float(r_old[64] or 0), float(r_old[65] or 0), float(r_old[66] or 0)),
            ("SNBT", "SNBT", 0.0, float(r_old[68] or 0), float(r_old[69] or 0), float(r_old[70] or 0)),
            ("SMMPTN", "SMMPTN Barat", 0.0, float(r_old[72] or 0), float(r_old[73] or 0), float(r_old[74] or 0)),
            ("TALENTA", "TALENTA USK", 0.0, float(r_old[76] or 0), float(r_old[77] or 0), float(r_old[78] or 0)),
            ("SMC", "SELEKSI MANDIRI CADANGAN", 0.0, float(r_old[80] or 0), float(r_old[81] or 0), float(r_old[82] or 0)),
            ("ADIK", "ADIK (Afirmasi)", 0.0, 0.0, float(r_old[84] or 0), float(r_old[85] or 0))
        ]
        for j_std, j_asli, pm, dt, la, du in j_2025_list:
            row_dict = make_jalur_row(2025, j_std, j_asli, pm, dt, la, du)
            jalur_records_2025.append(row_dict)
            jalur_records_all.append(row_dict)

        # 5. Jalur 2026 (SNBP, SNBT, SMMPTN, TALENTA, SMC, ADIK)
        dt_snbp_26 = float(r_26[19] or 0)
        pm_snbp_26 = float(r_26[22] or 0)
        la_snbp_26 = float(r_26[23] or 0)
        du_snbp_26 = float(r_26[33] or 0)

        dt_snbt_26 = float(r_26[36] or 0)
        pm_snbt_26 = float(r_26[37] or 0)
        la_snbt_26 = float(r_26[38] or 0)
        du_snbt_26 = float(r_26[42] or 0)

        dt_smm_26 = float(r_26[46] or 0)
        pm_smm_26 = float(r_26[48] or 0)
        la_smm_26 = float(r_26[49] or 0)
        du_smm_26 = float(r_26[53] or 0)

        pm_tal_26 = float(r_26[56] or 0)
        la_tal_26 = float(r_26[57] or 0)
        du_tal_26 = float(r_26[62] or 0)

        pm_smc_26 = float(r_26[64] or 0)
        la_smc_26 = float(r_26[65] or 0)
        du_smc_26 = float(r_26[67] or 0)

        la_adik_26 = float(r_26[74] or 0)
        du_adik_26 = float(r_26[76] or 0)

        j_2026_list = [
            ("SNBP", "SNBP", pm_snbp_26, dt_snbp_26, la_snbp_26, du_snbp_26),
            ("SNBT", "SNBT", pm_snbt_26, dt_snbt_26, la_snbt_26, du_snbt_26),
            ("SMMPTN", "SMMPTN Barat", pm_smm_26, dt_smm_26, la_smm_26, du_smm_26),
            ("TALENTA", "TALENTA USK", pm_tal_26, 0.0, la_tal_26, du_tal_26),
            ("SMC", "SELEKSI MANDIRI CADANGAN", pm_smc_26, 0.0, la_smc_26, du_smc_26),
            ("ADIK", "ADIK (Afirmasi)", 0.0, 0.0, la_adik_26, du_adik_26)
        ]
        for j_std, j_asli, pm, dt, la, du in j_2026_list:
            row_dict = make_jalur_row(2026, j_std, j_asli, pm, dt, la, du)
            jalur_records_2026.append(row_dict)
            jalur_records_all.append(row_dict)

        # Calculate annual metrics
        pem_vals = [pm_22, pm_23, pm_24, pm_25, pm_26]
        dt_vals = [dt_22, dt_23, dt_24, dt_25, dt_26]
        la_vals = [la_22, la_23, la_24, la_25, la_26]
        du_vals = [du_22, du_23, du_24, du_25, du_26]

        rec = {
            "No": no,
            "Fakultas": fakultas,
            "Program_Studi": nm_prodi,
            "Jenjang": jenjang,
            "Segmen_Analisis": klaster
        }

        for i, yr in enumerate([2022, 2023, 2024, 2025, 2026]):
            pm = pem_vals[i]
            dt = dt_vals[i]
            la = la_vals[i]
            du = du_vals[i]
            gugur = max(0.0, la - du)
            keketatan = round(pm / dt, 2) if dt > 0 else 0.0
            fill_rate = round(du / dt * 100.0, 1) if dt > 0 else 0.0
            yield_rate = round(du / la * 100.0, 1) if la > 0 else 0.0
            gap = max(0.0, dt - du)

            rec[f"Peminat_{yr}"] = int(pm)
            rec[f"DT_{yr}"] = int(dt)
            rec[f"LA_{yr}"] = int(la)
            rec[f"DU_{yr}"] = int(du)
            rec[f"Gugur_{yr}"] = int(gugur)
            rec[f"Keketatan_{yr}"] = keketatan
            rec[f"FillRate_{yr}_Persen"] = fill_rate
            rec[f"YieldRate_{yr}_Persen"] = yield_rate
            rec[f"Sisa_Kosong_{yr}"] = int(gap)

        # YoY Calculations for DU
        yoy_du_23 = round(((du_vals[1] - du_vals[0]) / du_vals[0] * 100.0), 1) if du_vals[0] > 0 else 0.0
        yoy_du_24 = round(((du_vals[2] - du_vals[1]) / du_vals[1] * 100.0), 1) if du_vals[1] > 0 else 0.0
        yoy_du_25 = round(((du_vals[3] - du_vals[2]) / du_vals[2] * 100.0), 1) if du_vals[2] > 0 else 0.0
        yoy_du_26 = round(((du_vals[4] - du_vals[3]) / du_vals[3] * 100.0), 1) if du_vals[3] > 0 else 0.0

        rec["YoY_DU_22_23"] = yoy_du_23
        rec["YoY_DU_23_24"] = yoy_du_24
        rec["YoY_DU_24_25"] = yoy_du_25
        rec["YoY_DU_25_26"] = yoy_du_26

        transitions_du = [yoy_du_23, yoy_du_24, yoy_du_25, yoy_du_26]
        valid_trans_du = [t for i, t in enumerate(transitions_du) if du_vals[i] > 0]
        naik_du = sum(1 for t in valid_trans_du if t > 0)
        turun_du = sum(1 for t in valid_trans_du if t < 0)
        rec["Riwayat_Transisi_DU"] = f"{naik_du}x Naik, {turun_du}x Turun"

        # YoY Calculations for Peminat
        yoy_pm_23 = round(((pem_vals[1] - pem_vals[0]) / pem_vals[0] * 100.0), 1) if pem_vals[0] > 0 else 0.0
        yoy_pm_24 = round(((pem_vals[2] - pem_vals[1]) / pem_vals[1] * 100.0), 1) if pem_vals[1] > 0 else 0.0
        yoy_pm_25 = round(((pem_vals[3] - pem_vals[2]) / pem_vals[2] * 100.0), 1) if pem_vals[2] > 0 else 0.0
        yoy_pm_26 = round(((pem_vals[4] - pem_vals[3]) / pem_vals[3] * 100.0), 1) if pem_vals[3] > 0 else 0.0

        rec["YoY_Peminat_22_23"] = yoy_pm_23
        rec["YoY_Peminat_23_24"] = yoy_pm_24
        rec["YoY_Peminat_24_25"] = yoy_pm_25
        rec["YoY_Peminat_25_26"] = yoy_pm_26

        # OLS Linear Regression Slopes (Orang/Tahun) & R2
        slope_du, r2_du = calculate_ols_slope(du_vals)
        slope_pm, r2_pm = calculate_ols_slope(pem_vals)
        rec["Slope_Tren_DU_Orang_Thn"] = slope_du
        rec["R2_Stabilitas_DU"] = r2_du
        rec["Slope_Tren_Peminat_Orang_Thn"] = slope_pm
        rec["R2_Stabilitas_Peminat"] = r2_pm

        # CAGR calculations
        rec["CAGR_DU_Persen"] = calculate_cagr(du_vals[0], du_vals[4], periods=4)
        rec["CAGR_DT_Persen"] = calculate_cagr(dt_vals[0], dt_vals[4], periods=4)
        rec["CAGR_Peminat_Persen"] = calculate_cagr(pem_vals[0], pem_vals[4], periods=4)

        # Marginal Fill Rate = (DU 2026 - DU 2022) / (DT 2026 - DT 2022)
        delta_du = du_vals[4] - du_vals[0]
        delta_dt = dt_vals[4] - dt_vals[0]
        if delta_dt > 0:
            rec["Marginal_Fill_Rate"] = round(delta_du / delta_dt, 3)
        elif delta_dt == 0:
            rec["Marginal_Fill_Rate"] = 1.0 if delta_du >= 0 else 0.0
        else:
            rec["Marginal_Fill_Rate"] = 0.0

        # 5-Year Averages
        rec["Rata_Peminat_5Thn"] = round(np.mean([p for p in pem_vals if p > 0]), 1) if any(p > 0 for p in pem_vals) else 0.0
        rec["Rata_DT_5Thn"] = round(np.mean([p for p in dt_vals if p > 0]), 1) if any(p > 0 for p in dt_vals) else 0.0
        rec["Rata_DU_5Thn"] = round(np.mean([p for p in du_vals if p > 0]), 1) if any(p > 0 for p in du_vals) else 0.0
        rec["Rata_FillRate_5Thn_Persen"] = round(np.mean([rec[f"FillRate_{yr}_Persen"] for yr in [2022, 2023, 2024, 2025, 2026] if rec[f"DT_{yr}"] > 0]), 1)
        rec["Rata_Keketatan_5Thn"] = round(np.mean([rec[f"Keketatan_{yr}"] for yr in [2022, 2023, 2024, 2025, 2026] if rec[f"DT_{yr}"] > 0]), 2)

        # 5 Official Classification Categories (Tahap 6 Panduan)
        if du_vals[0] == 0:
            kategori = "Data Terbatas (Prodi Baru)"
        elif klaster == "PSDKU Gayo Lues":
            kategori = "Peminatan Relatif Rendah (PSDKU)"
        elif klaster == "Diploma 3 Vokasi":
            kategori = "Diploma 3 Vokasi"
        elif rec["Rata_FillRate_5Thn_Persen"] < 68.0 or rec["Keketatan_2026"] < 1.5:
            if delta_dt > 30 and rec["Marginal_Fill_Rate"] < 0.35:
                kategori = "Over-Ekspansi Kuota (Defisit)"
            else:
                kategori = "Peminatan Relatif Rendah"
        elif (naik_du >= 3 or (not np.isnan(rec["CAGR_DU_Persen"]) and rec["CAGR_DU_Persen"] >= 10.0)) and slope_du > 5.0:
            kategori = "Tren Meningkat"
        elif (turun_du >= 3 or (not np.isnan(rec["CAGR_DU_Persen"]) and rec["CAGR_DU_Persen"] < -2.0)) and slope_du < -2.0:
            kategori = "Tren Menurun"
        elif abs(slope_du) <= 3.0 and rec["Rata_FillRate_5Thn_Persen"] >= 80.0:
            kategori = "Tren Stabil"
        else:
            kategori = "Tren Fluktuatif"

        rec["Kategori_Tren"] = kategori
        master_records.append(rec)

    df_master = pd.DataFrame(master_records)
    df_s1 = df_master[df_master["Segmen_Analisis"] == "S1 Kampus Utama"].copy()
    df_d3 = df_master[df_master["Segmen_Analisis"] == "Diploma 3 Vokasi"].copy()
    df_psdku = df_master[df_master["Segmen_Analisis"] == "PSDKU Gayo Lues"].copy()

    df_jalur_all = pd.DataFrame(jalur_records_all)
    df_j26 = pd.DataFrame(jalur_records_2026)
    df_j25 = pd.DataFrame(jalur_records_2025)
    df_j24 = pd.DataFrame(jalur_records_2024)
    df_j23 = pd.DataFrame(jalur_records_2023)
    df_j22 = pd.DataFrame(jalur_records_2022)

    print(f"Data Summary: S1={len(df_s1)}, D3={len(df_d3)}, PSDKU={len(df_psdku)}, Total={len(df_master)}")
    print(f"Jalur Summary: All={len(df_jalur_all)}, 2022={len(df_j22)}, 2023={len(df_j23)}, 2024={len(df_j24)}, 2025={len(df_j25)}, 2026={len(df_j26)}")

    # Write raw data
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df_master.to_excel(writer, sheet_name="Master_Semua_Prodi", index=False)
        df_s1.to_excel(writer, sheet_name="S1_Kampus_Utama", index=False)
        df_d3.to_excel(writer, sheet_name="Diploma_3_Vokasi", index=False)
        df_psdku.to_excel(writer, sheet_name="PSDKU_Gayo_Lues", index=False)
        df_jalur_all.to_excel(writer, sheet_name="Rincian_Jalur_Semua_Tahun", index=False)
        df_j26.to_excel(writer, sheet_name="Rincian_Jalur_Masuk_2026", index=False)
        df_j25.to_excel(writer, sheet_name="Rincian_Jalur_Masuk_2025", index=False)
        df_j24.to_excel(writer, sheet_name="Rincian_Jalur_Masuk_2024", index=False)
        df_j23.to_excel(writer, sheet_name="Rincian_Jalur_Masuk_2023", index=False)
        df_j22.to_excel(writer, sheet_name="Rincian_Jalur_Masuk_2022", index=False)

    print("Applying executive-level formatting, auto-fit column widths, and aesthetic styling...")
    format_master_workbook(out_file)
    print(f"Master file successfully created and beautified: {out_file}")

def format_master_workbook(file_path):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(file_path)

    HEADER_MAP = {
        "Tahun": "Tahun Akademik",
        "No": "No",
        "Fakultas": "Fakultas",
        "Program_Studi": "Nama Program Studi",
        "Jenjang": "Jenjang",
        "Segmen_Analisis": "Klaster Analisis",
        "Jalur_Penerimaan": "Jalur Masuk Standar",
        "Nama_Asli_Jalur": "Nama Jalur di Dokumen",
        
        # 2022
        "Peminat_2022": "Peminat 2022",
        "DT_2022": "Daya Tampung 2022",
        "LA_2022": "Lulus Seleksi 2022",
        "DU_2022": "Daftar Ulang 2022",
        "Gugur_2022": "Mundur/Gugur 2022",
        "Keketatan_2022": "Keketatan 2022",
        "FillRate_2022_Persen": "Fill Rate 2022 (%)",
        "YieldRate_2022_Persen": "Yield Rate 2022 (%)",
        "Sisa_Kosong_2022": "Kursi Kosong 2022",

        # 2023
        "Peminat_2023": "Peminat 2023",
        "DT_2023": "Daya Tampung 2023",
        "LA_2023": "Lulus Seleksi 2023",
        "DU_2023": "Daftar Ulang 2023",
        "Gugur_2023": "Mundur/Gugur 2023",
        "Keketatan_2023": "Keketatan 2023",
        "FillRate_2023_Persen": "Fill Rate 2023 (%)",
        "YieldRate_2023_Persen": "Yield Rate 2023 (%)",
        "Sisa_Kosong_2023": "Kursi Kosong 2023",

        # 2024
        "Peminat_2024": "Peminat 2024",
        "DT_2024": "Daya Tampung 2024",
        "LA_2024": "Lulus Seleksi 2024",
        "DU_2024": "Daftar Ulang 2024",
        "Gugur_2024": "Mundur/Gugur 2024",
        "Keketatan_2024": "Keketatan 2024",
        "FillRate_2024_Persen": "Fill Rate 2024 (%)",
        "YieldRate_2024_Persen": "Yield Rate 2024 (%)",
        "Sisa_Kosong_2024": "Kursi Kosong 2024",

        # 2025
        "Peminat_2025": "Peminat 2025",
        "DT_2025": "Daya Tampung 2025",
        "LA_2025": "Lulus Seleksi 2025",
        "DU_2025": "Daftar Ulang 2025",
        "Gugur_2025": "Mundur/Gugur 2025",
        "Keketatan_2025": "Keketatan 2025",
        "FillRate_2025_Persen": "Fill Rate 2025 (%)",
        "YieldRate_2025_Persen": "Yield Rate 2025 (%)",
        "Sisa_Kosong_2025": "Kursi Kosong 2025",

        # 2026
        "Peminat_2026": "Peminat 2026",
        "DT_2026": "Daya Tampung 2026",
        "LA_2026": "Lulus Seleksi 2026",
        "DU_2026": "Daftar Ulang 2026",
        "Gugur_2026": "Mundur/Gugur 2026",
        "Keketatan_2026": "Keketatan 2026",
        "FillRate_2026_Persen": "Fill Rate 2026 (%)",
        "YieldRate_2026_Persen": "Yield Rate 2026 (%)",
        "Sisa_Kosong_2026": "Kursi Kosong 2026",

        # YoY Transitions
        "YoY_DU_22_23": "YoY DU 22-23 (%)",
        "YoY_DU_23_24": "YoY DU 23-24 (%)",
        "YoY_DU_24_25": "YoY DU 24-25 (%)",
        "YoY_DU_25_26": "YoY DU 25-26 (%)",
        "Riwayat_Transisi_DU": "Riwayat Transisi DU",

        "YoY_Peminat_22_23": "YoY Peminat 22-23 (%)",
        "YoY_Peminat_23_24": "YoY Peminat 23-24 (%)",
        "YoY_Peminat_24_25": "YoY Peminat 24-25 (%)",
        "YoY_Peminat_25_26": "YoY Peminat 25-26 (%)",

        # Advanced Slopes & Metrics
        "Slope_Tren_DU_Orang_Thn": "Slope DU (Orang/Thn)",
        "R2_Stabilitas_DU": "R² Stabilitas DU",
        "Slope_Tren_Peminat_Orang_Thn": "Slope Peminat (Orang/Thn)",
        "R2_Stabilitas_Peminat": "R² Stabilitas Peminat",
        "CAGR_DU_Persen": "CAGR DU (%)",
        "CAGR_DT_Persen": "CAGR DT (%)",
        "CAGR_Peminat_Persen": "CAGR Peminat (%)",
        "Marginal_Fill_Rate": "Marginal Fill Rate",

        # 5-Year Averages
        "Rata_Peminat_5Thn": "Rata Peminat 5-Thn",
        "Rata_DT_5Thn": "Rata DT 5-Thn",
        "Rata_DU_5Thn": "Rata DU 5-Thn",
        "Rata_FillRate_5Thn_Persen": "Rata Fill Rate 5-Thn (%)",
        "Rata_Keketatan_5Thn": "Rata Keketatan 5-Thn",
        "Kategori_Tren": "Klasifikasi Tren Resmi",

        # Rincian Jalur 2026
        "Jalur_Penerimaan": "Jalur Penerimaan",
        "Peminat": "Jumlah Peminat",
        "Daya_Tampung": "Target Daya Tampung",
        "Lulus_Seleksi": "Calon Lulus Seleksi",
        "Daftar_Ulang": "Mahasiswa Daftar Ulang",
        "Tidak_Daftar_Ulang": "Mundur / Gugur",
        "Yield_Rate_Persen": "Yield Rate (%)"
    }

    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC')
    )

    # Color definitions for headers by theme
    FILL_BASE = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Deep USK Navy
    FILL_2022 = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Slate
    FILL_2023 = PatternFill(start_color="1F618D", end_color="1F618D", fill_type="solid") # Ocean
    FILL_2024 = PatternFill(start_color="2874A6", end_color="2874A6", fill_type="solid") # Steel
    FILL_2025 = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid") # Sky
    FILL_2026 = PatternFill(start_color="117864", end_color="117864", fill_type="solid") # Deep Emerald (Highlight 2026)
    FILL_TREN = PatternFill(start_color="6C3483", end_color="6C3483", fill_type="solid") # Executive Purple
    FILL_RATA = PatternFill(start_color="7D6608", end_color="7D6608", fill_type="solid") # Dark Amber Gold

    FONT_HEADER = Font(name="Segoe UI", size=10.5, bold=True, color="FFFFFF")
    FONT_DATA = Font(name="Segoe UI", size=9.5, color="1C2833")
    FONT_DATA_BOLD = Font(name="Segoe UI", size=9.5, bold=True, color="1C2833")

    FILL_ZEBRA_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    FILL_ZEBRA_ODD = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
    ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.views.sheetView[0].showGridLines = True

        # Set row heights
        ws.row_dimensions[1].height = 28.0

        # Translate Headers and apply styling
        num_cols = ws.max_column
        num_rows = ws.max_row

        raw_headers = []
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            raw_val = str(cell.value or '').strip()
            raw_headers.append(raw_val)
            nice_val = HEADER_MAP.get(raw_val, raw_val.replace('_', ' '))
            cell.value = nice_val
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_HEADER
            cell.border = thin_border

            # Apply thematic header color
            if "2022" in sheetname:
                cell.fill = FILL_2022
            elif "2023" in sheetname:
                cell.fill = FILL_2023
            elif "2024" in sheetname:
                cell.fill = FILL_2024
            elif "2025" in sheetname:
                cell.fill = FILL_2025
            elif "2026" in sheetname:
                cell.fill = FILL_2026
            elif any(k in raw_val for k in ["2022"]):
                cell.fill = FILL_2022
            elif any(k in raw_val for k in ["2023"]):
                cell.fill = FILL_2023
            elif any(k in raw_val for k in ["2024"]):
                cell.fill = FILL_2024
            elif any(k in raw_val for k in ["2025"]):
                cell.fill = FILL_2025
            elif any(k in raw_val for k in ["2026"]):
                cell.fill = FILL_2026
            elif any(k in raw_val for k in ["YoY", "Slope", "R2", "CAGR", "Marginal", "Transisi"]):
                cell.fill = FILL_TREN
            elif any(k in raw_val for k in ["Rata", "Kategori"]):
                cell.fill = FILL_RATA
            else:
                cell.fill = FILL_BASE

        # Freeze panes: freeze columns A..C (No, Fakultas, Program Studi) or A..D (if Tahun exists)
        if "Tahun" in raw_headers:
            ws.freeze_panes = "E2"
        else:
            ws.freeze_panes = "D2"

        # Enable AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # Format data cells
        for row_idx in range(2, num_rows + 1):
            ws.row_dimensions[row_idx].height = 20.0
            row_fill = FILL_ZEBRA_EVEN if (row_idx % 2 == 0) else FILL_ZEBRA_ODD

            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                raw_h = raw_headers[col_idx - 1]
                val = cell.value

                cell.border = thin_border
                cell.fill = row_fill
                cell.font = FONT_DATA

                # Formatting based on column semantics
                if raw_h in ["No", "Jenjang", "Tahun"]:
                    cell.alignment = ALIGN_CENTER
                elif raw_h in ["Fakultas", "Program_Studi", "Segmen_Analisis", "Jalur_Penerimaan", "Nama_Asli_Jalur", "Riwayat_Transisi_DU"]:
                    cell.alignment = ALIGN_LEFT
                    if raw_h == "Program_Studi":
                        cell.font = FONT_DATA_BOLD
                elif raw_h == "Kategori_Tren":
                    cell.alignment = ALIGN_LEFT
                    cell.font = FONT_DATA_BOLD
                    # Subtle highlight for critical prodi
                    if "Over-Ekspansi" in str(val) or "Rendah" in str(val):
                        cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="922B21")
                    elif "Meningkat" in str(val):
                        cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="117864")
                    elif "Menurun" in str(val):
                        cell.font = Font(name="Segoe UI", size=9.5, bold=True, color="B7950B")
                elif any(k in raw_h for k in ["Persen", "YoY", "Rate", "CAGR"]):
                    cell.alignment = ALIGN_RIGHT
                    if isinstance(val, (int, float)):
                        cell.number_format = '0.0"%"'
                elif any(k in raw_h for k in ["Keketatan", "Slope", "R2", "Marginal"]):
                    cell.alignment = ALIGN_RIGHT
                    if isinstance(val, (int, float)):
                        cell.number_format = '0.00'
                else: # Default numeric (counts)
                    cell.alignment = ALIGN_RIGHT
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0'

        # Auto-adjust column width with extra breathing room
        for col_idx in range(1, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            raw_h = raw_headers[col_idx - 1]
            nice_h = str(ws.cell(row=1, column=col_idx).value or '')

            max_len = len(nice_h)
            for row_idx in range(2, num_rows + 1):
                val_str = str(ws.cell(row=row_idx, column=col_idx).value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)

            col_width = max(max_len + 4, 12)
            if "Program" in nice_h:
                col_width = max(col_width, 42)
            elif "Fakultas" in nice_h:
                col_width = max(col_width, 25)
            elif "Klasifikasi" in nice_h or "Kategori" in nice_h:
                col_width = max(col_width, 32)
            elif "Dokumen" in nice_h or "Asli" in nice_h:
                col_width = max(col_width, 30)
            elif "Standar" in nice_h or "Jalur" in nice_h:
                col_width = max(col_width, 22)
            elif "Klaster" in nice_h:
                col_width = max(col_width, 20)
            elif "Transisi" in nice_h:
                col_width = max(col_width, 24)
            elif "Tahun" in nice_h:
                col_width = max(col_width, 16)

            ws.column_dimensions[col_letter].width = col_width

    wb.save(file_path)

if __name__ == "__main__":
    main()

